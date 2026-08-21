# =============================================================
#  AGRO-SYNC - Backend v2
#  Agora com banco SQLite, append-only e protecao contra duplicata
# =============================================================

import io
import os
import sqlite3
from datetime import datetime, timezone

from flask import Flask, request, jsonify, render_template, g, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Caminho do arquivo do banco. E so um arquivo na pasta do projeto.
BANCO = os.path.join(os.path.dirname(__file__), "agro.db")


# -------------------------------------------------------------
# 1) CONEXAO COM O BANCO
# -------------------------------------------------------------
# g e uma "mochila" que o Flask cria a cada requisicao e joga fora
# no fim. Guardamos a conexao ali para nao abrir uma nova a cada
# consulta dentro da mesma requisicao.
def conectar():
    if "db" not in g:
        g.db = sqlite3.connect(BANCO)
        g.db.row_factory = sqlite3.Row   # devolve linhas tipo dicionario
    return g.db


@app.teardown_appcontext
def fechar(excecao):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# -------------------------------------------------------------
# 2) CRIACAO DA TABELA
# -------------------------------------------------------------
# Note o que NAO existe aqui: nenhuma coluna e alterada depois de
# gravada. Nao ha UPDATE nem DELETE em lugar nenhum deste arquivo.
#
#   id_local   -> UUID gerado no CELULAR. E a chave contra duplicata.
#   corrige_id -> se este registro corrige outro, aponta para o antigo.
#                 O errado continua no banco, visivel. Isso e a trilha.
def criar_tabela():
    with sqlite3.connect(BANCO) as db:
        db.execute("""
            CREATE TABLE IF NOT EXISTS aplicacoes (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                id_local    TEXT    NOT NULL UNIQUE,
                produto     TEXT    NOT NULL,
                quantidade  TEXT    NOT NULL,
                talhao      TEXT    NOT NULL,
                criado_em   TEXT    NOT NULL,
                recebido_em TEXT    NOT NULL,
                corrige_id  TEXT
            )
        """)


# -------------------------------------------------------------
# 3) TELA
# -------------------------------------------------------------
@app.route("/")
def home():
    return render_template("index.html")


# -------------------------------------------------------------
# 4) SINCRONIZAR
# -------------------------------------------------------------
@app.route("/api/sincronizar", methods=["POST"])
def sincronizar():
    dados = request.get_json(silent=True)

    if not dados or "registros" not in dados:
        return jsonify({"status": "erro",
                        "mensagem": "Envie um JSON com a chave 'registros'."}), 400

    db = conectar()
    gravados = 0
    repetidos = 0
    recusados = []

    for r in dados["registros"]:

        # 4.1) VALIDACAO NO SERVIDOR.
        #      A validacao do JavaScript ajuda o usuario honesto.
        #      Esta aqui e a unica que vale de verdade.
        id_local   = (r.get("id_local")   or "").strip()
        produto    = (r.get("produto")    or "").strip()
        quantidade = (r.get("quantidade") or "").strip()
        talhao     = (r.get("talhao")     or "").strip()

        if not (id_local and produto and quantidade and talhao):
            recusados.append({"id_local": id_local or "?",
                              "motivo": "campo obrigatorio vazio"})
            continue

        # 4.2) GRAVACAO IDEMPOTENTE.
        #      Se a rede caiu DEPOIS que o servidor gravou mas ANTES da
        #      resposta chegar, o celular vai reenviar tudo. Sem essa
        #      protecao, o registro entraria duas vezes.
        #      O UNIQUE em id_local faz o SQLite recusar a copia, e o
        #      "OR IGNORE" faz ele recusar em silencio, sem erro.
        cursor = db.execute("""
            INSERT OR IGNORE INTO aplicacoes
                (id_local, produto, quantidade, talhao,
                 criado_em, recebido_em, corrige_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            id_local, produto, quantidade, talhao,
            r.get("criado_em", ""),
            datetime.now(timezone.utc).isoformat(timespec="seconds"),
            r.get("corrige_id")
        ))

        if cursor.rowcount == 1:
            gravados += 1
        else:
            repetidos += 1   # ja estava no banco. Nao e erro.

    db.commit()

    # 4.3) Log no terminal, para voce enxergar o que aconteceu.
    print(f"\n[{datetime.now():%H:%M:%S}] SINCRONIZACAO"
          f" | novos: {gravados}"
          f" | repetidos: {repetidos}"
          f" | recusados: {len(recusados)}")
    for x in recusados:
        print(f"   recusado {x['id_local']}: {x['motivo']}")

    return jsonify({
        "status": "ok",
        "gravados": gravados,
        "repetidos": repetidos,
        "recusados": recusados,
        "mensagem": f"{gravados} novo(s), {repetidos} ja existia(m)."
    }), 200


# -------------------------------------------------------------
# 5) CONSULTAR / EXPORTAR REGISTROS
# -------------------------------------------------------------
def _parse_data(dia: str | None) -> str | None:
    """Aceita YYYY-MM-DD. Devolve None se invalido."""
    if not dia:
        return None
    try:
        datetime.strptime(dia.strip()[:10], "%Y-%m-%d")
        return dia.strip()[:10]
    except ValueError:
        return None


def buscar_aplicacoes(de: str | None = None, ate: str | None = None):
    db = conectar()
    sql = "SELECT * FROM aplicacoes WHERE 1=1"
    params: list[str] = []

    if de:
        sql += " AND criado_em >= ?"
        params.append(f"{de}T00:00:00")

    if ate:
        sql += " AND criado_em <= ?"
        params.append(f"{ate}T23:59:59")

    sql += " ORDER BY criado_em ASC"
    return db.execute(sql, params).fetchall()


def _formatar_data(iso: str) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return iso


def _nome_arquivo(de: str | None, ate: str | None) -> str:
    if de and ate:
        return f"agro-sync_{de}_a_{ate}.xlsx"
    if de:
        return f"agro-sync_desde_{de}.xlsx"
    if ate:
        return f"agro-sync_ate_{ate}.xlsx"
    return "agro-sync_todas-aplicacoes.xlsx"


def _texto_periodo(de: str | None, ate: str | None) -> str:
    if de and ate:
        return f"Período: {de[8:10]}/{de[5:7]}/{de[0:4]} a {ate[8:10]}/{ate[5:7]}/{ate[0:4]}"
    if de:
        return f"A partir de {de[8:10]}/{de[5:7]}/{de[0:4]}"
    if ate:
        return f"Até {ate[8:10]}/{ate[5:7]}/{ate[0:4]}"
    return "Período: todas as aplicações"


def _mesclar_registros(do_servidor: list[dict], pendentes: list[dict]) -> list[dict]:
    ids_servidor = {r["id_local"] for r in do_servidor}
    extras = []

    for r in pendentes:
        id_local = (r.get("id_local") or "").strip()
        if not id_local or id_local in ids_servidor:
            continue
        status = "Aguardando envio"
        if r.get("erro"):
            status = f"Erro: {r['erro']}"
        extras.append({
            "produto": r.get("produto", ""),
            "quantidade": r.get("quantidade", ""),
            "talhao": r.get("talhao", ""),
            "criado_em": r.get("criado_em", ""),
            "recebido_em": "",
            "status": status,
        })

    registros = []
    for r in do_servidor:
        registros.append({
            "produto": r["produto"],
            "quantidade": r["quantidade"],
            "talhao": r["talhao"],
            "criado_em": r["criado_em"],
            "recebido_em": r.get("recebido_em", ""),
            "status": "Sincronizado",
        })
    registros.extend(extras)
    registros.sort(key=lambda x: x.get("criado_em") or "")
    return registros


def gerar_planilha_xlsx(registros: list[dict], de: str | None, ate: str | None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Aplicações"
    ws.sheet_view.showGridLines = False

    colunas = [
        ("Produto", 28),
        ("Quantidade", 16),
        ("Talhão", 18),
        ("Data da aplicação", 22),
        ("Enviado em", 22),
        ("Status", 20),
    ]
    ultima_col = len(colunas)

    verde = "2F6B3A"
    verde_claro = "E2EFE4"
    papel = "F2F0E9"
    branco = "FFFFFF"
    ambar_claro = "F6E9D5"
    vermelho_claro = "FDECEA"

    fino = Side(style="thin", color="D3CFC2")
    borda = Border(left=fino, right=fino, top=fino, bottom=fino)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_col)
    titulo = ws.cell(row=1, column=1, value="Agro-Sync — Relatório de Aplicações")
    titulo.font = Font(name="Segoe UI", size=16, bold=True, color=verde)
    titulo.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtítulo (período)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_col)
    subtitulo = ws.cell(row=2, column=1, value=_texto_periodo(de, ate))
    subtitulo.font = Font(name="Segoe UI", size=11, color="6C6D63")
    subtitulo.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6

    linha_cabecalho = 4
    cabecalho_fill = PatternFill("solid", fgColor=verde)
    cabecalho_font = Font(name="Segoe UI", size=11, bold=True, color=branco)

    for idx, (nome, largura) in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=idx, value=nome)
        celula.fill = cabecalho_fill
        celula.font = cabecalho_font
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda
        ws.column_dimensions[get_column_letter(idx)].width = largura

    ws.row_dimensions[linha_cabecalho].height = 26
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{linha_cabecalho}:{get_column_letter(ultima_col)}{linha_cabecalho}"

    dados_font = Font(name="Segoe UI", size=11, color="17211B")
    alt_fill = PatternFill("solid", fgColor=papel)

    status_sync_fill = PatternFill("solid", fgColor=verde_claro)
    status_sync_font = Font(name="Segoe UI", size=11, bold=True, color=verde)
    status_erro_fill = PatternFill("solid", fgColor=vermelho_claro)
    status_erro_font = Font(name="Segoe UI", size=11, bold=True, color="B3261E")
    status_pendente_fill = PatternFill("solid", fgColor=ambar_claro)
    status_pendente_font = Font(name="Segoe UI", size=11, bold=True, color="B06A12")

    for i, reg in enumerate(registros):
        linha = linha_cabecalho + 1 + i
        valores = [
            reg["produto"],
            reg["quantidade"],
            reg["talhao"],
            _formatar_data(reg["criado_em"]),
            _formatar_data(reg["recebido_em"]) if reg.get("recebido_em") else "—",
            reg["status"],
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha, column=col, value=valor)
            celula.font = dados_font
            celula.border = borda
            celula.alignment = Alignment(
                horizontal="left" if col != 6 else "center",
                vertical="center",
                wrap_text=(col == 1),
            )
            if i % 2 == 1:
                celula.fill = alt_fill

        status_cel = ws.cell(row=linha, column=6)
        status_txt = reg["status"]
        if status_txt == "Sincronizado":
            status_cel.fill = status_sync_fill
            status_cel.font = status_sync_font
        elif status_txt.startswith("Erro"):
            status_cel.fill = status_erro_fill
            status_cel.font = status_erro_font
        else:
            status_cel.fill = status_pendente_fill
            status_cel.font = status_pendente_font

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@app.route("/api/registros", methods=["GET"])
def registros():
    de = _parse_data(request.args.get("de"))
    ate = _parse_data(request.args.get("ate"))

    if request.args.get("de") and de is None:
        return jsonify({"status": "erro", "mensagem": "Data inicial invalida."}), 400
    if request.args.get("ate") and ate is None:
        return jsonify({"status": "erro", "mensagem": "Data final invalida."}), 400

    linhas = buscar_aplicacoes(de, ate)
    return jsonify([dict(l) for l in linhas]), 200


@app.route("/api/exportar", methods=["POST"])
def exportar():
    dados = request.get_json(silent=True) or {}
    de = _parse_data(dados.get("de"))
    ate = _parse_data(dados.get("ate"))

    if dados.get("de") and de is None:
        return jsonify({"status": "erro", "mensagem": "Data inicial invalida."}), 400
    if dados.get("ate") and ate is None:
        return jsonify({"status": "erro", "mensagem": "Data final invalida."}), 400

    do_servidor = [dict(l) for l in buscar_aplicacoes(de, ate)]
    pendentes = dados.get("pendentes") or []
    registros = _mesclar_registros(do_servidor, pendentes)

    if not registros:
        return jsonify({"status": "erro", "mensagem": "Nenhuma aplicacao encontrada."}), 404

    xlsx = gerar_planilha_xlsx(registros, de, ate)
    nome = _nome_arquivo(de, ate)

    return Response(
        xlsx,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


if __name__ == "__main__":
    criar_tabela()
    app.run(debug=True, port=5000)
